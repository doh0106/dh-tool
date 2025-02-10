from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from .utils import map_column_names_to_letters, apply_to_cells

COLOR_MAP = {
    "black": "000000",
    "white": "FFFFFF",
    "red": "FF0000",
    "green": "00FF00",
    "blue": "0000FF",
    "yellow": "FFFF00",
    "cyan": "00FFFF",
    "magenta": "FF00FF",
    "gray": "808080",
    "orange": "FFA500",
    "purple": "800080",
    "pink": "FFC0CB",
    "brown": "A52A2A",
    "gold": "FFD700",
    "silver": "C0C0C0",
}


class Style:

    @staticmethod
    def apply_auto_wrap(worksheet):
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    @staticmethod
    def freeze_first_row(worksheet):
        worksheet.freeze_panes = worksheet["A2"]

    @staticmethod
    def set_column_width(worksheet, width_map):
        """컬럼 이름 또는 엑셀 열 문자로 열 너비 설정"""
        col_letter_map = map_column_names_to_letters(worksheet, width_map)
        for col, width in col_letter_map.items():
            worksheet.column_dimensions[col].width = width

    @staticmethod
    def auto_adjust_column_widths(worksheet):
        """데이터에 맞게 자동으로 열 너비 조정"""
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    @staticmethod
    def apply_auto_filter(worksheet, columns=None):
        """
        데이터 필터 적용
        - columns: None이면 모든 컬럼에 필터 적용
                   리스트면 특정 컬럼에만 필터 적용
        """
        max_row = worksheet.max_row

        if columns is None:  # ✅ 모든 열에 필터 적용
            max_col = worksheet.max_column
            worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        elif isinstance(columns, list):  # ✅ 특정 열에만 필터 적용
            col_indices = []
            header = [cell.value for cell in worksheet[1]]
            for col in columns:
                if col in header:
                    col_idx = header.index(col) + 1
                    col_indices.append(col_idx)

            if not col_indices:
                raise ValueError("지정한 컬럼이 시트에 존재하지 않습니다.")

            col_letters = [get_column_letter(idx) for idx in col_indices]
            ref_range = f"{col_letters[0]}1:{col_letters[-1]}{max_row}"
            worksheet.auto_filter.ref = ref_range

    @staticmethod
    @apply_to_cells
    def set_font(cell, font_name="Arial", font_size=12, bold=False, italic=False):
        font = Font(name=font_name, size=font_size, bold=bold, italic=italic)
        cell.font = font
        # for row in worksheet.iter_rows(min_row=1, max_row=1):
        #     for cell in row:
        #         cell.font = font

    @staticmethod
    @apply_to_cells
    def apply_border(cell, border_style="thin"):
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style),
        )
        cell.border = border
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         cell.border = border

    @staticmethod
    @apply_to_cells
    def apply_color(cell, color):
        """
        셀 배경색 적용
        - 16진수 색상 코드("FFFF00") 또는 색상 이름("red") 지원
        """
        # ✅ 색상 이름을 16진수로 변환
        if color.lower() in COLOR_MAP:
            color = COLOR_MAP[color.lower()]

        # ✅ 16진수 형식 보정
        if not color.startswith("#") and len(color) == 6:
            color = f"FF{color}"  # openpyxl은 ARGB 포맷을 사용

        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = fill
        # for row in worksheet.iter_rows():
        #     for cell in row:
        #         if cell.value:
        #             cell.fill = fill


MY_COLOR_MAP = {
    "pastel_blue": {
        "name": "Pastel Blue",
        "hex": "#E3F2FD",
        "argb": "FFE3F2FD",
        "category": "blue",
    },
    "soft_blue": {
        "name": "Soft Blue",
        "hex": "#BBDEFB",
        "argb": "FFBBDEFB",
        "category": "blue",
    },
    "deep_blue": {
        "name": "Deep Blue",
        "hex": "#90CAF9",
        "argb": "FF90CAF9",
        "category": "blue",
    },
    "pastel_pink": {
        "name": "Pastel Pink",
        "hex": "#FCE4EC",
        "argb": "FFFCE4EC",
        "category": "pink",
    },
    "soft_pink": {
        "name": "Soft Pink",
        "hex": "#F8BBD0",
        "argb": "FFF8BBD0",
        "category": "pink",
    },
    "deep_pink": {
        "name": "Deep Pink",
        "hex": "#F48FB1",
        "argb": "FFF48FB1",
        "category": "pink",
    },
    "pastel_mint": {
        "name": "Pastel Mint",
        "hex": "#E8F5E9",
        "argb": "FFE8F5E9",
        "category": "mint",
    },
    "soft_mint": {
        "name": "Soft Mint",
        "hex": "#C8E6C9",
        "argb": "FFC8E6C9",
        "category": "mint",
    },
    "deep_mint": {
        "name": "Deep Mint",
        "hex": "#A5D6A7",
        "argb": "FFA5D6A7",
        "category": "mint",
    },
    "pastel_yellow": {
        "name": "Pastel Yellow",
        "hex": "#FFFDE7",
        "argb": "FFFFFDE7",
        "category": "yellow",
    },
    "soft_yellow": {
        "name": "Soft Yellow",
        "hex": "#FFF9C4",
        "argb": "FFFFF9C4",
        "category": "yellow",
    },
    "deep_yellow": {
        "name": "Deep Yellow",
        "hex": "#FFF59D",
        "argb": "FFFFF59D",
        "category": "yellow",
    },
    "pastel_purple": {
        "name": "Pastel Purple",
        "hex": "#F3E5F5",
        "argb": "FFF3E5F5",
        "category": "purple",
    },
    "soft_purple": {
        "name": "Soft Purple",
        "hex": "#E1BEE7",
        "argb": "FFE1BEE7",
        "category": "purple",
    },
    "deep_purple": {
        "name": "Deep Purple",
        "hex": "#CE93D8",
        "argb": "FFCE93D8",
        "category": "purple",
    },
}
