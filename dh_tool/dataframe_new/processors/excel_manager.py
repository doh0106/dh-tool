import json
import numpy as np
import pandas as pd
from datetime import datetime
import warnings

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


class ExcelExporter:
    """엑셀로 데이터 저장 및 변환"""

    @staticmethod
    def create_workbook():
        """새로운 엑셀 워크북 생성"""
        return Workbook()

    @staticmethod
    def sync_df_to_worksheet(worksheet, dataframe: pd.DataFrame):
        """DataFrame을 엑셀 워크시트에 반영"""
        worksheet.delete_rows(1, worksheet.max_row)
        for row in dataframe_to_rows(dataframe, index=False, header=True):
            worksheet.append(row)

    @staticmethod
    def save(workbook: Workbook, filename: str):
        """엑셀 파일 저장"""
        workbook.save(filename)


class ExcelSheetManager:
    """엑셀 시트 관리"""

    workbook = Workbook()  # 클래스 변수로 워크북 관리
    sheet_names = {"Sheet1"}  # 시트 이름을 추적하기 위한 클래스 변수

    @classmethod
    def create_sheet(cls, dataframe, title="Sheet"):
        """새로운 시트를 추가"""
        if title in cls.sheet_names:
            print(f"{title} sheet은 이미 존재합니다, 이름을 바꿔주세요")
            return False
        cls.sheet_names.add(title)
        cls.workbook.create_sheet(title=title)
        return True

    @classmethod
    def select_sheet(cls, title):
        """특정 시트를 선택"""
        if title not in cls.workbook.sheetnames:
            raise ValueError(f"{title} 시트가 존재하지 않습니다.")
        return cls.workbook[title]

    @classmethod
    def list_sheets(cls):
        """워크북 내 모든 시트 목록 반환"""
        return cls.workbook.sheetnames


class ExcelStyler:
    """엑셀 스타일 관리"""

    @staticmethod
    def apply_style(worksheet, style_func):
        """사용자 정의 스타일 적용"""
        for row in worksheet.iter_rows():
            for cell in row:
                style_func(cell)

    @staticmethod
    def color_cells(worksheet, condition, color):
        """조건에 맞는 셀에 색상 적용"""
        for row in worksheet.iter_rows():
            for cell in row:
                if condition(cell.value):
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    @staticmethod
    def auto_adjust_columns(worksheet):
        """열 너비 자동 조정"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    @staticmethod
    def enable_autowrap(worksheet):
        """자동 줄바꿈 적용"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    @staticmethod
    def set_column_width(df, worksheet, **kwargs):
        """컬럼 너비 설정"""
        for column, width in kwargs.items():
            if column in df.columns:
                col_idx = df.columns.get_loc(column) + 1
                col_letter = chr(64 + col_idx)
                worksheet.column_dimensions[col_letter].width = width
            else:
                pass


class ExcelManager:
    def __init__(self, dataframe: pd.DataFrame, add_column_prefix=False):
        self.df = dataframe
        self.workbook = ExcelSheetManager.workbook
        self.worksheet = self.workbook.active
        self.worksheet.title = "Sheet1"
        self.sync_df_to_worksheet()

    def sync_df_to_worksheet(self):
        self.worksheet.delete_rows(1, self.worksheet.max_row)

        def _convert_to_string_if_needed(value):
            """복잡한 데이터 타입을 문자열로 변환"""
            try:
                if isinstance(value, (dict, list, np.ndarray)):
                    warnings.warn(
                        f"Complex data type {type(value)} will be converted to string representation"
                    )
                    return str(value)  # json.dumps 대신 str 사용
                elif pd.isna(value):
                    return ""
                elif isinstance(value, (np.int64, np.float64)):
                    return float(value)
                elif isinstance(value, datetime):
                    return value.isoformat()
                elif value in (np.inf, -np.inf):
                    return str(value)
                return value
            except Exception as e:
                warnings.warn(
                    f"Error converting value {value}: {str(e)}. Using string representation."
                )
                return str(value)

        # if self.add_column_prefix:
        #     # 열 이름에 접두사 추가 (이미 'col_'로 시작하는 경우 제외)
        #     self.df.columns = [
        #         col if col.startswith("col_") else f"col_{col}"
        #         for col in self.df.columns
        #     ]

        for row in dataframe_to_rows(self.df, index=False, header=True):
            row = [_convert_to_string_if_needed(cell) for cell in row]
            self.worksheet.append(row)

    @property
    def df(self):
        return self._df

    @df.setter
    def df(self, dataframe):
        self._df = dataframe

    def update(self, df):
        self.df = df
        self.sync_df_to_worksheet()

    def create_sheet(self, title="Sheet"):
        return ExcelSheetManager.create_sheet(self.workbook, title)

    def select_sheet(self, title):
        return ExcelSheetManager.select_sheet(self.workbook, title)

    def list_sheets(self):
        return ExcelSheetManager.list_sheets(self.workbook)

    def save(self, filename):
        ExcelExporter.save(self.workbook, filename)

    def create_sheet(self, title="Sheet") -> bool:
        return ExcelSheetManager.create_sheet(self.workbook, title)

    def select_sheet(self, title):
        return ExcelSheetManager.select_sheet(self.workbook, title)

    def list_sheets(self):
        return ExcelSheetManager.list_sheets(self.workbook)

    def auto_adjust_columns(self):
        ExcelStyler.auto_adjust_columns(self.worksheet)

    def enable_autowrap(self):
        ExcelStyler.enable_autowrap(self.worksheet)

    def color_cells(self, condition, color):
        ExcelStyler.color_cells(self.worksheet, condition, color)

    def set_columns_width(self, **kwargs):
        ExcelStyler.set_column_width(self.df, self.worksheet, **kwargs)
